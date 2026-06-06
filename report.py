"""
report.py — Generate portfolio_report.xlsx after each run.
4 sheets: Positions, Historique, Scores, Performance.
"""

import logging
from datetime import datetime, date

import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

logger = logging.getLogger(__name__)

# ─── Style helpers ────────────────────────────────────────────────────────────

_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
_GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
_RED_FILL   = PatternFill("solid", fgColor="FFC7CE")
_ALT_FILL   = PatternFill("solid", fgColor="EEF2FF")
_HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
_BOLD       = Font(bold=True, size=10)
_THIN       = Side(style="thin", color="CCCCCC")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_row(ws, cols, row=1):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _BORDER


def _autowidth(ws, min_w=10, max_w=30):
    for col in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, length + 2))


def _pct_fill(cell, value):
    """Apply green/red fill based on sign of value."""
    if value is None:
        return
    cell.fill = _GREEN_FILL if value >= 0 else _RED_FILL


def _fetch_prices(tickers):
    prices = {}
    for t in tickers:
        try:
            hist = yf.Ticker(t).history(period="5d")
            if not hist.empty:
                prices[t] = float(hist["Close"].iloc[-1])
        except Exception:
            pass
    return prices


# ─── Sheet builders ───────────────────────────────────────────────────────────

def _build_positions(wb, state):
    ws = wb.active
    ws.title = "Positions"

    positions = state.get("positions", {})
    regime    = state.get("current_regime", "?")
    cash      = state.get("cash_eur", 0)
    initial   = state.get("initial_capital", 1890)

    # Fetch live prices for open positions
    live_prices = _fetch_prices(list(positions.keys())) if positions else {}

    cols = [
        "Ticker", "Bourse", "Entrée (€)", "Prix actuel (€)", "Nb titres",
        "Valeur (€)", "P&L (€)", "P&L (%)", "Stop-loss (€)", "Dist. Stop (%)",
        "Take-profit (€)", "Dist. TP (%)", "Beta", "Entrée date", "Jours"
    ]
    _header_row(ws, cols)

    def _exchange(ticker):
        if ticker.endswith(".PA"): return "Euronext Paris"
        if ticker.endswith(".AS"): return "Euronext Amsterdam"
        if ticker.endswith(".MI"): return "Borsa Italiana"
        return "Euronext"

    today = date.today()
    total_pnl = 0.0

    for r, (ticker, pos) in enumerate(positions.items(), 2):
        entry   = pos.get("entry_price", 0)
        current = live_prices.get(ticker, entry)
        shares  = pos.get("nb_shares", 0)
        stop    = pos.get("stop_loss", 0)
        tp      = pos.get("take_profit", 0)
        beta    = pos.get("beta", 1.0)
        val     = current * shares
        pnl_eur = (current - entry) * shares
        pnl_pct = (current - entry) / entry if entry else 0
        dist_stop = (current - stop) / current if current else 0
        dist_tp   = (tp - current) / current if current else 0
        total_pnl += pnl_eur

        try:
            entry_d  = date.fromisoformat(pos.get("entry_date", str(today)))
            days     = (today - entry_d).days
        except Exception:
            days = 0

        row_data = [
            ticker, _exchange(ticker), entry, round(current, 2), shares,
            round(val, 2), round(pnl_eur, 2), pnl_pct, stop, dist_stop,
            tp, dist_tp, beta, pos.get("entry_date", ""), days
        ]
        fill = _ALT_FILL if r % 2 == 0 else None
        for c, val_ in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val_)
            cell.border    = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill

        # Color P&L cells
        _pct_fill(ws.cell(row=r, column=7), pnl_eur)
        _pct_fill(ws.cell(row=r, column=8), pnl_pct)

        # Format pct columns
        ws.cell(row=r, column=8).number_format  = "0.00%"
        ws.cell(row=r, column=10).number_format = "0.00%"
        ws.cell(row=r, column=12).number_format = "0.00%"

    # Summary row
    sr = len(positions) + 3
    ws.cell(row=sr, column=1, value="RÉSUMÉ").font = _BOLD
    ws.cell(row=sr, column=2, value=f"Régime: {regime}")
    ws.cell(row=sr, column=3, value=f"Cash: {cash:.0f} €")
    ws.cell(row=sr, column=4, value=f"P&L total: {total_pnl:+.2f} €").font = _BOLD
    ws.cell(row=sr, column=5, value=f"Capital initial: {initial:.0f} €")

    _autowidth(ws)
    ws.freeze_panes = "A2"


def _build_historique(wb, state):
    ws = wb.create_sheet("Historique")
    history = state.get("trade_history", [])

    cols = ["Date sortie", "Ticker", "Nb titres", "Prix entrée (€)",
            "Prix sortie (€)", "P&L (€)", "P&L (%)", "Date entrée"]
    _header_row(ws, cols)

    for r, trade in enumerate(reversed(history), 2):
        row_data = [
            trade.get("exit_date", ""),
            trade.get("ticker", ""),
            trade.get("nb_shares", 0),
            trade.get("entry_price", 0),
            trade.get("exit_price", 0),
            trade.get("pnl_eur", 0),
            trade.get("pnl_pct", 0),
            trade.get("entry_date", ""),
        ]
        fill = _ALT_FILL if r % 2 == 0 else None
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill

        pnl_eur = trade.get("pnl_eur", 0)
        pnl_pct = trade.get("pnl_pct", 0)
        _pct_fill(ws.cell(row=r, column=6), pnl_eur)
        _pct_fill(ws.cell(row=r, column=7), pnl_pct)
        ws.cell(row=r, column=7).number_format = "0.00%"

    if not history:
        ws.cell(row=2, column=1, value="Aucun trade clôturé pour l'instant.")

    _autowidth(ws)
    ws.freeze_panes = "A2"


def _build_scores(wb, state):
    ws = wb.create_sheet("Scores")
    last_scores = state.get("last_scores", [])

    cols = [
        "Ticker", "Score", "Tech", "Fund", "Bonus",
        "Trend", "RSI", "Volume", "MACD", "Momentum", "Bollinger", "StochRSI",
        "ATR%", "Beta", "Régime"
    ]
    _header_row(ws, cols)

    regime = state.get("current_regime", "?")

    for r, s in enumerate(last_scores, 2):
        bd = s.get("breakdown", {})
        row_data = [
            s.get("ticker", ""),
            round(s.get("score", 0), 1),
            round(s.get("tech_score", 0), 1),
            round(s.get("fund_score", 0), 1),
            round(s.get("regime_bonus", 0), 1),
            round(bd.get("trend", 0), 1),
            round(bd.get("rsi", 0), 1),
            round(bd.get("volume", 0), 1),
            round(bd.get("macd", 0), 1),
            round(bd.get("momentum", 0), 1),
            round(bd.get("bollinger", 0), 1),
            round(bd.get("stoch_rsi", 0), 1),
            s.get("atr_pct"),
            round(s.get("beta", 1.0), 2),
            regime,
        ]
        fill = _ALT_FILL if r % 2 == 0 else None
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border    = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill

        # Highlight top 5
        score = s.get("score", 0)
        if r <= 6:
            ws.cell(row=r, column=1).font = Font(bold=True, color="1F3864")
            ws.cell(row=r, column=2).fill = _GREEN_FILL

        if s.get("atr_pct") is not None:
            ws.cell(row=r, column=13).number_format = "0.00%"

    if not last_scores:
        ws.cell(row=2, column=1, value="Aucun score disponible (lancer un run d'abord).")

    _autowidth(ws)
    ws.freeze_panes = "A2"


def _build_performance(wb, state):
    ws = wb.create_sheet("Performance")
    perf    = state.get("performance", {})
    initial = state.get("initial_capital", 1890)
    cash    = state.get("cash_eur", 0)
    pnl_eur = perf.get("total_pnl_eur", 0)
    pnl_pct = perf.get("total_pnl_pct", 0)

    positions   = state.get("positions", {})
    live_prices = _fetch_prices(list(positions.keys())) if positions else {}
    pos_val     = sum(live_prices.get(t, p.get("entry_price", 0)) * p.get("nb_shares", 0)
                      for t, p in positions.items())
    total_val   = pos_val + cash

    # KPI cards
    kpi_rows = [
        ("Capital initial",     f"{initial:.2f} €"),
        ("Valeur du portefeuille", f"{total_val:.2f} €"),
        ("Cash disponible",     f"{cash:.2f} €"),
        ("P&L total",           f"{pnl_eur:+.2f} €"),
        ("P&L (%)",             f"{pnl_pct:+.2%}"),
        ("Positions ouvertes",  len(positions)),
        ("Régime actuel",       state.get("current_regime", "?")),
        ("Dernier run",         state.get("last_run", "N/A")),
    ]

    ws.cell(row=1, column=1, value="INDICATEURS CLÉ").font = Font(bold=True, size=12, color="1F3864")
    for r, (label, value) in enumerate(kpi_rows, 2):
        ws.cell(row=r, column=1, value=label).font  = _BOLD
        cell = ws.cell(row=r, column=2, value=value)
        cell.border = _BORDER
        if "P&L" in label and isinstance(value, str):
            cell.fill = _GREEN_FILL if pnl_eur >= 0 else _RED_FILL

    # Weekly returns history table
    weekly = perf.get("weekly_returns", [])
    if weekly:
        ws.cell(row=12, column=1, value="Historique hebdomadaire").font = _BOLD
        ws.cell(row=13, column=1, value="Semaine").font = _HDR_FONT
        ws.cell(row=13, column=1).fill = _HDR_FILL
        ws.cell(row=13, column=2, value="Rendement").font = _HDR_FONT
        ws.cell(row=13, column=2).fill = _HDR_FILL

        for r, ret in enumerate(weekly, 14):
            ws.cell(row=r, column=1, value=r - 13)
            cell = ws.cell(row=r, column=2, value=ret)
            cell.number_format = "0.00%"
            _pct_fill(cell, ret)

        # Chart
        if len(weekly) >= 2:
            chart = LineChart()
            chart.title  = "Rendements hebdomadaires"
            chart.y_axis.title = "Rendement"
            chart.x_axis.title = "Semaine"
            data = Reference(ws, min_col=2, min_row=13, max_row=13 + len(weekly))
            chart.add_data(data, titles_from_data=True)
            chart.width  = 20
            chart.height = 12
            ws.add_chart(chart, "D2")

    _autowidth(ws)


# ─── Entry point ──────────────────────────────────────────────────────────────

def generate_report(state, filename="portfolio_report.xlsx"):
    try:
        wb = Workbook()
        _build_positions(wb, state)
        _build_historique(wb, state)
        _build_scores(wb, state)
        _build_performance(wb, state)
        wb.save(filename)
        logger.info(f"Report saved: {filename}")
    except Exception as e:
        logger.error(f"Report generation failed: {e}")
