"""
backtest.py — Combined Model V1 Final:
- CORE (60%): Simple 9M momentum rotation, rebalance every 42 days
- SATELLITE (40%): Daily swing, offensive in BULL/NEUTRAL, defensive in BEAR
- Target: 20%+ CAGR, DD < 20%, regular trades
"""
import json
import logging
from datetime import datetime
from itertools import product
import pandas as pd
import numpy as np
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import LineChart, Reference

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

START_DATE   = "2020-01-01"
END_DATE     = "2025-12-31"
INITIAL_CASH = 10000.0
FEE_US       = 2.00
CORE_PCT      = 0.60
SATELLITE_PCT = 0.40

YF_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
    "Referer": "https://finance.yahoo.com",
}

CORE_UNIVERSE = [
    "AAPL", "MSFT", "NVDA", "AMZN", "GOOGL", "META", "TSLA", "AMD",
    "AVGO", "ORCL", "CRM", "ADBE", "NOW", "PANW", "SNPS", "CDNS",
    "LMT", "RTX", "NOC", "GD",
    "LLY", "ABBV", "ISRG", "DXCM",
    "V", "MA", "GS", "MS", "JPM",
    "COST", "HD", "WMT", "PG",
    "XOM", "CVX",
    "QQQ", "XLK", "XLE", "XLF", "XLV", "XLI", "XLB", "XLP",
]

SATELLITE_OFFENSIVE = [
    # Leveraged ETFs
    "TQQQ", "SOXL", "UPRO", "TECL", "LABU", "FAS", "TNA", "SPXL",
    # Crypto-adjacent
    "COIN", "MSTR", "RIOT", "MARA", "CLSK", "HUT", "BITF",
    # High beta tech
    "PLTR", "SMCI", "IONQ", "RBLX", "HOOD", "SOFI", "AFRM",
    "UPST", "HIMS", "SOUN",
    # AI plays
    "NVDA", "AMD", "ARM", "MRVL",
    # Biotech
    "MRNA", "BNTX",
    # High momentum / space
    "DKNG", "RKLB", "ASTS", "TSLA", "RIVN",
    # ETFs volatils
    "XBI", "ARKK",
]

SATELLITE_DEFENSIVE = []

ALL_TICKERS = list(set(
    CORE_UNIVERSE + SATELLITE_OFFENSIVE + SATELLITE_DEFENSIVE
))

PARAM_GRID = {
    "core_n":           [5, 8],
    "core_mom_days":    [189],
    "core_ma":          [150, 200],
    "rebalance_days":   [42],
    "sat_score_thresh": [33, 38],
    "sat_stop_atr":     [2.0, 2.5],
    "sat_take_profit":  [0.22, 0.28],
    "sat_hold_days":    [20, 35],
}


def fetch_history(ticker):
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
        r = requests.get(url, headers=YF_HEADERS,
                         params={"interval": "1d", "range": "7y"}, timeout=15)
        if r.status_code != 200:
            return None
        data   = r.json()
        result = data.get("chart", {}).get("result")
        if not result:
            return None
        ts    = result[0]["timestamp"]
        quote = result[0]["indicators"]["quote"][0]
        dates = [datetime.utcfromtimestamp(t).date() for t in ts]
        df = pd.DataFrame({
            "date":   dates,
            "close":  quote.get("close", []),
            "high":   quote.get("high", []),
            "low":    quote.get("low", []),
            "volume": quote.get("volume", []),
        }).dropna().set_index("date")
        return df if len(df) > 252 else None
    except Exception as e:
        logger.warning(f"{ticker}: {e}")
        return None


def calc_momentum(closes, days):
    if len(closes) < days + 21:
        return None
    mom = (closes[21] - closes[days]) / closes[days]
    if len(closes) >= 63:
        rets = [(closes[j]-closes[j+1])/closes[j+1] for j in range(62)]
        vol  = (sum(r**2 for r in rets)/len(rets))**0.5 * (252**0.5)
        return mom / vol if vol > 0 else mom
    return mom


def calc_atr(highs, lows, closes, period=14):
    if len(highs) < period + 1:
        return 0.02
    tr_list = [max(highs[i]-lows[i],
                   abs(highs[i]-closes[i+1]),
                   abs(lows[i]-closes[i+1])) for i in range(period)]
    return sum(tr_list) / period / closes[0] if closes[0] > 0 else 0.02


def calc_rsi(closes, period=14):
    if len(closes) < period + 1:
        return 50.0
    deltas = [closes[i] - closes[i+1] for i in range(period)]
    gains  = sum(d for d in deltas if d > 0) / period
    losses = sum(-d for d in deltas if d < 0) / period
    rs     = gains / losses if losses != 0 else 100
    return 100 - 100 / (1 + rs)


def calc_ema(data, span):
    k = 2/(span+1); e = data[-1]
    for p in reversed(data[:-1]): e = p*k + e*(1-k)
    return e


def score_satellite(closes, highs, lows, volumes, regime):
    if len(closes) < 50:
        return 0.0, 0.02

    score = 0.0
    ma20  = sum(closes[:20]) / 20
    ma50  = sum(closes[:50]) / 50
    tp    = 0.0

    if len(closes) >= 200:
        ma100 = sum(closes[:100]) / 100
        ma200 = sum(closes[:200]) / 200
        if closes[0] > ma20:       tp += 3.0
        if ma20 > ma50:            tp += 3.0
        if ma50 > ma100:           tp += 3.0
        if ma100 > ma200:          tp += 3.0
        ma50_prev = sum(closes[5:55])/50 if len(closes) >= 55 else ma50
        if ma50 > ma50_prev:       tp += 1.0
        if closes[0] > ma200*1.02: tp += 1.0
    elif len(closes) >= 100:
        ma100 = sum(closes[:100]) / 100
        if closes[0] > ma20: tp += 2.5
        if ma20 > ma50:      tp += 2.5
        if ma50 > ma100:     tp += 3.0
        tp *= 0.85
    if regime == "BEAR": tp *= 0.7
    score += min(tp, 12.0)

    rsi = calc_rsi(closes)
    if 40 <= rsi <= 62:  score += 8.0
    elif rsi < 25:       score += 6.0
    elif rsi < 40:       score += 4.0
    elif rsi > 75:       score += 0.0
    else:                score += 3.0

    if len(volumes) >= 21:
        avg_vol = sum(volumes[1:21]) / 20
        avg5    = sum(volumes[:5]) / 5
        ratio   = volumes[0] / avg_vol if avg_vol > 0 else 0
        vt      = avg5 / avg_vol if avg_vol > 0 else 1.0
        if ratio >= 1.5 and vt > 1.1: score += 8.0
        elif ratio >= 1.5:             score += 6.0
        elif ratio >= 1.2:             score += 4.8
        elif ratio >= 0.8:             score += 2.4

    if len(closes) >= 29:
        macd   = calc_ema(closes[:12],12) - calc_ema(closes[:26],26)
        sig    = calc_ema(closes[:9],9)
        hist   = macd - sig
        macd_p = calc_ema(closes[3:15],12) - calc_ema(closes[3:29],26)
        sig_p  = calc_ema(closes[3:12],9)
        hist_p = macd_p - sig_p
        if hist > 0 and hist_p <= 0:     score += 6.0
        elif hist > 0 and macd > sig:    score += 4.0
        elif hist > hist_p and hist > 0: score += 3.0
        elif macd > 0:                   score += 1.5

    mp = 0.0
    if len(closes) >= 21:
        r1m = (closes[0]-closes[20])/closes[20]
        mp += 2.0 if r1m > 0.03 else (1.0 if r1m > 0 else 0.0)
    if len(closes) >= 63:
        r3m = (closes[0]-closes[62])/closes[62]
        mp += 3.0 if r3m > 0.05 else (1.5 if r3m > 0 else 0.0)
    if len(closes) >= 126:
        r6m = (closes[0]-closes[125])/closes[125]
        mp += 3.0 if r6m > 0.08 else (1.5 if r6m > 0 else 0.0)
    if regime == "BEAR": mp *= 0.6
    score += min(mp, 8.0)

    sma = sum(closes[:20])/20
    std = (sum((x-sma)**2 for x in closes[:20])/20)**0.5
    if std > 0:
        pctb = (closes[0]-(sma-2*std))/(4*std)
        if regime == "BEAR":
            if 0.1 <= pctb <= 0.45:  score += 7.0
            elif pctb < 0.1:          score += 4.0
            elif pctb <= 0.65:        score += 2.0
        else:
            if pctb >= 0.6:           score += 5.0
            elif pctb >= 0.4:         score += 3.0
            elif 0.1 <= pctb < 0.4:   score += 5.0

    stoch = 3.6 if 20 <= rsi <= 60 else (0.0 if rsi > 80 else 2.4)
    score += stoch

    atr_pct = calc_atr(highs, lows, closes) if highs and lows else 0.02
    if regime == "BEAR" and atr_pct > 0.04:
        score *= 0.85

    score += 15.0
    return min(100.0, score), atr_pct


def detect_regime(bench_closes, ma_period=200):
    if len(bench_closes) < ma_period:
        return "NEUTRAL"
    ma50  = sum(bench_closes[:50]) / 50
    ma200 = sum(bench_closes[:ma_period]) / ma_period
    mom   = (bench_closes[0]-bench_closes[99])/bench_closes[99] if len(bench_closes) >= 100 else 0
    if ma50 > ma200*1.02 and mom > 0:      return "BULL"
    elif ma50 < ma200*0.98 or mom < -0.05: return "BEAR"
    return "NEUTRAL"


def calc_annual_perf(trades, all_dates, equity):
    years = {}
    for i, d in enumerate(all_dates):
        yr = d.year
        if yr not in years:
            years[yr] = {"start_eq": equity[i], "end_eq": equity[i],
                         "trades": 0, "wins": 0}
        years[yr]["end_eq"] = equity[i]
    for t in trades:
        yr = int(t["entry_date"][:4])
        if yr in years:
            years[yr]["trades"] += 1
            if t["pnl_pct"] > 0: years[yr]["wins"] += 1
    results = []
    for yr in sorted(years.keys()):
        y   = years[yr]
        ret = (y["end_eq"]/y["start_eq"]-1)*100 if y["start_eq"] > 0 else 0
        wr  = y["wins"]/y["trades"]*100 if y["trades"] > 0 else 0
        results.append({
            "year": yr, "return": round(ret,1),
            "trades": y["trades"], "win_rate": round(wr,1),
        })
    return results


def run_single(all_data, bench_df, all_dates, params):
    core_n        = params["core_n"]
    core_mom_days = params["core_mom_days"]
    core_ma       = params["core_ma"]
    rebal_days    = params["rebalance_days"]
    sat_thresh    = params["sat_score_thresh"]
    sat_stop_atr  = params["sat_stop_atr"]
    sat_tp        = params["sat_take_profit"]
    sat_hold      = params["sat_hold_days"]

    core_cash     = INITIAL_CASH * CORE_PCT
    sat_cash      = INITIAL_CASH * SATELLITE_PCT
    core_holdings = {}
    sat_holdings  = {}
    trades        = []
    equity        = []
    peak_eq       = INITIAL_CASH
    last_rebal    = None

    bench_list = list(bench_df.index)
    min_idx    = max(core_mom_days + 21, core_ma, 200)

    for i, today in enumerate(all_dates):
        total_val = core_cash + sat_cash
        for ticker, pos in {**core_holdings, **sat_holdings}.items():
            if ticker in all_data and today in all_data[ticker].index:
                total_val += pos["shares"] * all_data[ticker].loc[today,"close"]
        equity.append(total_val)
        if total_val > peak_eq:
            peak_eq = total_val

        if i < min_idx:
            continue

        b_idx = bench_list.index(today) if today in bench_list else -1
        if b_idx < core_ma:
            regime  = "NEUTRAL"
            in_bear = False
        else:
            b_closes = bench_df["close"].iloc[b_idx-max(core_ma,200):b_idx+1].tolist()[::-1]
            regime   = detect_regime(b_closes, core_ma)
            ma_val   = sum(b_closes[:core_ma]) / core_ma
            in_bear  = b_closes[0] < ma_val

        # ── CORE: Bear exit ───────────────────────────────────────────────
        if in_bear and core_holdings:
            for ticker in list(core_holdings.keys()):
                if ticker in all_data and today in all_data[ticker].index:
                    pos       = core_holdings[ticker]
                    cur_price = all_data[ticker].loc[today,"close"]
                    pnl       = (cur_price-pos["entry_price"])/pos["entry_price"]
                    core_cash += pos["shares"]*cur_price - FEE_US
                    trades.append({
                        "ticker":     ticker,
                        "entry_date": str(pos["entry_date"]),
                        "exit_date":  str(today),
                        "entry_price":round(pos["entry_price"],4),
                        "exit_price": round(cur_price,4),
                        "pnl_pct":    round(pnl*100,2),
                        "reason":     "bear_exit",
                        "days_held":  (today-pos["entry_date"]).days,
                        "layer":      "CORE",
                    })
            core_holdings.clear()
            last_rebal = today

        # ── CORE: Rebalance every 42 days ─────────────────────────────────
        do_rebal = (last_rebal is None or (today-last_rebal).days >= rebal_days)

        if not in_bear and do_rebal:
            scores = []
            for ticker in CORE_UNIVERSE:
                if ticker not in all_data or today not in all_data[ticker].index:
                    continue
                t_idx  = list(all_data[ticker].index).index(today)
                closes = all_data[ticker]["close"].iloc[max(0,t_idx-core_mom_days-30):t_idx+1].tolist()[::-1]
                if len(closes) >= 200:
                    ma200 = sum(closes[:200])/200
                    if closes[0] < ma200*0.95:
                        continue
                mom = calc_momentum(closes, core_mom_days)
                if mom is not None:
                    scores.append((ticker, mom, all_data[ticker].loc[today,"close"]))

            scores.sort(key=lambda x: x[1], reverse=True)
            target = [s[0] for s in scores[:core_n]]

            for ticker in list(core_holdings.keys()):
                if ticker not in target:
                    if ticker in all_data and today in all_data[ticker].index:
                        pos       = core_holdings[ticker]
                        cur_price = all_data[ticker].loc[today,"close"]
                        pnl       = (cur_price-pos["entry_price"])/pos["entry_price"]
                        core_cash += pos["shares"]*cur_price - FEE_US
                        trades.append({
                            "ticker":     ticker,
                            "entry_date": str(pos["entry_date"]),
                            "exit_date":  str(today),
                            "entry_price":round(pos["entry_price"],4),
                            "exit_price": round(cur_price,4),
                            "pnl_pct":    round(pnl*100,2),
                            "reason":     "rotation",
                            "days_held":  (today-pos["entry_date"]).days,
                            "layer":      "CORE",
                        })
                    del core_holdings[ticker]

            core_total = core_cash
            for ticker, pos in core_holdings.items():
                if ticker in all_data and today in all_data[ticker].index:
                    core_total += pos["shares"]*all_data[ticker].loc[today,"close"]

            for ticker in target:
                if ticker in core_holdings:
                    continue
                if ticker not in all_data or today not in all_data[ticker].index:
                    continue
                price     = all_data[ticker].loc[today,"close"]
                slot_size = core_total / core_n
                invest    = min(slot_size, core_cash*0.95)
                if invest < price: continue
                shares = int(invest/price)
                cost   = shares*price + FEE_US
                if cost <= core_cash and shares > 0:
                    core_cash -= cost
                    core_holdings[ticker] = {
                        "shares":      shares,
                        "entry_price": price,
                        "entry_date":  today,
                    }
            last_rebal = today

        # ── CORE: Daily stop-loss -7% ─────────────────────────────────────
        for ticker in list(core_holdings.keys()):
            if ticker not in all_data or today not in all_data[ticker].index:
                continue
            pos       = core_holdings[ticker]
            cur_price = all_data[ticker].loc[today,"close"]
            pnl       = (cur_price-pos["entry_price"])/pos["entry_price"]
            if pnl <= -0.07:
                core_cash += pos["shares"]*cur_price - FEE_US
                trades.append({
                    "ticker":     ticker,
                    "entry_date": str(pos["entry_date"]),
                    "exit_date":  str(today),
                    "entry_price":round(pos["entry_price"],4),
                    "exit_price": round(cur_price,4),
                    "pnl_pct":    round(pnl*100,2),
                    "reason":     "stop_loss",
                    "days_held":  (today-pos["entry_date"]).days,
                    "layer":      "CORE",
                })
                del core_holdings[ticker]

        # ── SATELLITE: Daily exits ────────────────────────────────────────
        for ticker in list(sat_holdings.keys()):
            if ticker not in all_data or today not in all_data[ticker].index:
                continue
            pos        = sat_holdings[ticker]
            cur_price  = all_data[ticker].loc[today,"close"]
            pnl        = (cur_price-pos["entry_price"])/pos["entry_price"]
            days_held  = (today-pos["entry_date"]).days
            fixed_stop = -pos["atr_pct"] * sat_stop_atr

            sell = False; reason = ""
            if pnl <= fixed_stop:        sell = True; reason = "stop_loss"
            elif pnl >= sat_tp:          sell = True; reason = "take_profit"
            elif days_held >= sat_hold:  sell = True; reason = "timeout"

            if sell:
                sat_cash += pos["shares"]*cur_price - FEE_US
                trades.append({
                    "ticker":     ticker,
                    "entry_date": str(pos["entry_date"]),
                    "exit_date":  str(today),
                    "entry_price":round(pos["entry_price"],4),
                    "exit_price": round(cur_price,4),
                    "pnl_pct":    round(pnl*100,2),
                    "reason":     reason,
                    "days_held":  days_held,
                    "layer":      "SATELLITE",
                    "regime":     regime,
                })
                del sat_holdings[ticker]

        # ── SATELLITE: Daily entries ──────────────────────────────────────
        sat_universe = SATELLITE_OFFENSIVE if not in_bear else []

        max_sat = 4
        if len(sat_holdings) < max_sat and sat_cash > 0:
            sat_scores = []
            for ticker in sat_universe:
                if ticker in sat_holdings or ticker not in all_data:
                    continue
                if today not in all_data[ticker].index:
                    continue
                t_idx   = list(all_data[ticker].index).index(today)
                closes  = all_data[ticker]["close"].iloc[max(0,t_idx-250):t_idx+1].tolist()[::-1]
                highs   = all_data[ticker]["high"].iloc[max(0,t_idx-250):t_idx+1].tolist()[::-1] if "high" in all_data[ticker].columns else []
                lows    = all_data[ticker]["low"].iloc[max(0,t_idx-250):t_idx+1].tolist()[::-1]  if "low"  in all_data[ticker].columns else []
                volumes = all_data[ticker]["volume"].iloc[max(0,t_idx-250):t_idx+1].tolist()[::-1]

                score, atr_pct = score_satellite(closes, highs, lows, volumes, regime)
                if score >= sat_thresh:
                    sat_scores.append((ticker, score, all_data[ticker].loc[today,"close"], atr_pct))

            sat_scores.sort(key=lambda x: x[1], reverse=True)
            slots = max_sat - len(sat_holdings)
            for ticker, score, price, atr_pct in sat_scores[:slots]:
                slot_size = (INITIAL_CASH*SATELLITE_PCT) / max_sat
                invest    = min(slot_size, sat_cash*0.95)
                if invest < price: continue
                shares = int(invest/price)
                cost   = shares*price + FEE_US
                if cost <= sat_cash and shares > 0:
                    sat_cash -= cost
                    sat_holdings[ticker] = {
                        "shares":      shares,
                        "entry_price": price,
                        "entry_date":  today,
                        "atr_pct":     atr_pct,
                    }

    # Close all at end
    last_date = all_dates[-1]
    for ticker, pos in {**core_holdings, **sat_holdings}.items():
        if ticker in all_data and last_date in all_data[ticker].index:
            cur_price = all_data[ticker].loc[last_date,"close"]
            pnl = (cur_price-pos["entry_price"])/pos["entry_price"]
            layer = "CORE" if ticker in core_holdings else "SATELLITE"
            if layer == "CORE": core_cash += pos["shares"]*cur_price - FEE_US
            else:               sat_cash  += pos["shares"]*cur_price - FEE_US
            trades.append({
                "ticker":     ticker,
                "entry_date": str(pos["entry_date"]),
                "exit_date":  str(last_date),
                "entry_price":round(pos["entry_price"],4),
                "exit_price": round(cur_price,4),
                "pnl_pct":    round(pnl*100,2),
                "reason":     "end_of_backtest",
                "days_held":  (last_date-pos["entry_date"]).days,
                "layer":      layer,
            })

    if not trades:
        return None

    final_cash = core_cash + sat_cash
    total_ret  = (final_cash/INITIAL_CASH-1)*100
    years      = len(all_dates)/252
    cagr       = ((final_cash/INITIAL_CASH)**(1/years)-1)*100

    wins     = [t for t in trades if t["pnl_pct"] > 0]
    losses   = [t for t in trades if t["pnl_pct"] <= 0]
    win_rate = len(wins)/len(trades)*100 if trades else 0
    avg_win  = sum(t["pnl_pct"] for t in wins)/len(wins) if wins else 0
    avg_loss = sum(t["pnl_pct"] for t in losses)/len(losses) if losses else 0
    pf       = abs(avg_win*len(wins)/(avg_loss*len(losses))) if losses and avg_loss else 0

    eq_s   = pd.Series(equity)
    peak   = eq_s.cummax()
    max_dd = float(((eq_s-peak)/peak).min()*100)
    dr     = eq_s.pct_change().dropna()
    sharpe = float((dr.mean()/dr.std()*np.sqrt(252))) if dr.std() > 0 else 0

    core_trades = [t for t in trades if t.get("layer") == "CORE"]
    sat_trades  = [t for t in trades if t.get("layer") == "SATELLITE"]
    annual      = calc_annual_perf(trades, all_dates, equity)

    return {
        "params":         params,
        "total_return":   round(total_ret,1),
        "cagr":           round(cagr,1),
        "win_rate":       round(win_rate,1),
        "avg_win":        round(avg_win,1),
        "avg_loss":       round(avg_loss,1),
        "profit_factor":  round(pf,2),
        "max_drawdown":   round(max_dd,1),
        "sharpe":         round(sharpe,2),
        "n_trades":       len(trades),
        "n_core_trades":  len(core_trades),
        "n_sat_trades":   len(sat_trades),
        "equity":         [round(e,2) for e in equity],
        "trades":         trades,
        "annual":         annual,
    }


def run_backtest():
    logger.info("Downloading data...")
    all_data = {}
    for t in ALL_TICKERS:
        df = fetch_history(t)
        if df is not None:
            all_data[t] = df
            logger.info(f"  {t}: {len(df)} days")
        else:
            logger.warning(f"  {t}: skip")

    bench_df = fetch_history("SPY")
    if bench_df is None:
        logger.error("Failed SPY"); return

    bench_ret = (bench_df["close"].iloc[-1]/bench_df["close"].iloc[0]-1)*100
    all_dates = sorted(d for d in bench_df.index
                       if pd.to_datetime(START_DATE).date() <= d <=
                       pd.to_datetime(END_DATE).date())

    n = len(list(product(*PARAM_GRID.values())))
    logger.info(f"SPY: {bench_ret:.1f}% | {len(all_dates)} days | {n} combos")

    results = []
    for combo in product(*PARAM_GRID.values()):
        params = dict(zip(PARAM_GRID.keys(), combo))
        r = run_single(all_data, bench_df, all_dates, params)
        if r:
            results.append(r)
            logger.info(
                f"  core_n={params['core_n']} mom={params['core_mom_days']}d "
                f"ma={params['core_ma']} rebal={params['rebalance_days']}d "
                f"sat={params['sat_score_thresh']} "
                f"atr={params['sat_stop_atr']}x tp={params['sat_take_profit']} "
                f"hold={params['sat_hold_days']}d "
                f"→ {r['total_return']}% CAGR={r['cagr']}% "
                f"DD={r['max_drawdown']}% Sharpe={r['sharpe']} "
                f"trades={r['n_trades']} (core={r['n_core_trades']} sat={r['n_sat_trades']})"
            )

    if not results:
        logger.error("No results"); return

    best_sharpe = max(results, key=lambda x: x["sharpe"])
    best_return = max(results, key=lambda x: x["total_return"])

    logger.info(f"\n{'='*60}")
    logger.info(f"BEST (Sharpe): {best_sharpe['params']}")
    logger.info(f"  Return={best_sharpe['total_return']}% CAGR={best_sharpe['cagr']}% "
                f"DD={best_sharpe['max_drawdown']}% Sharpe={best_sharpe['sharpe']} "
                f"trades={best_sharpe['n_trades']} "
                f"(core={best_sharpe['n_core_trades']} sat={best_sharpe['n_sat_trades']})")
    logger.info(f"BEST (Return): {best_return['params']}")
    logger.info(f"  Return={best_return['total_return']}% CAGR={best_return['cagr']}% "
                f"DD={best_return['max_drawdown']}% Sharpe={best_return['sharpe']} "
                f"trades={best_return['n_trades']} "
                f"(core={best_return['n_core_trades']} sat={best_return['n_sat_trades']})")
    logger.info(f"SPY: {bench_ret:.1f}%")
    logger.info("Annual performance (best Sharpe):")
    for y in best_sharpe.get("annual", []):
        logger.info(f"  {y['year']}: {y['return']:+.1f}% | "
                    f"{y['trades']} trades | WR={y['win_rate']:.0f}%")
    logger.info(f"{'='*60}")

    with open("backtest_trades.json", "w") as f:
        json.dump(best_sharpe["trades"], f, indent=2)

    # ── Excel ──
    wb  = openpyxl.Workbook()
    hf  = PatternFill("solid", fgColor="1a1a2e")
    hft = Font(color="FFFFFF", bold=True)
    gf  = PatternFill("solid", fgColor="c8e6c9")
    rf  = PatternFill("solid", fgColor="ffcdd2")
    yf2 = PatternFill("solid", fgColor="fff9c4")
    bf  = PatternFill("solid", fgColor="bbdefb")

    ws1 = wb.active
    ws1.title = "Optimization"
    hdrs = ["Core N","Mom Days","MA","Rebal","Sat Thresh","Sat ATR",
            "Sat TP","Sat Hold","Total%","CAGR%","WR%","DD%","Sharpe",
            "Trades","Core","Sat"]
    for c, h in enumerate(hdrs, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hft
    for r_idx, r in enumerate(sorted(results, key=lambda x: x["sharpe"], reverse=True), 2):
        p = r["params"]
        vals = [p["core_n"], p["core_mom_days"], p["core_ma"], p["rebalance_days"],
                p["sat_score_thresh"], p["sat_stop_atr"], p["sat_take_profit"],
                p["sat_hold_days"], r["total_return"], r["cagr"], r["win_rate"],
                r["max_drawdown"], r["sharpe"], r["n_trades"],
                r["n_core_trades"], r["n_sat_trades"]]
        fill = gf if r["total_return"] > 150 else (
               bf if r["total_return"] > 80 else (
               yf2 if r["total_return"] > 0 else rf))
        for c, v in enumerate(vals, 1):
            ws1.cell(row=r_idx, column=c, value=v).fill = fill
    for col in "ABCDEFGHIJKLMNOP":
        ws1.column_dimensions[col].width = 11

    ws2 = wb.create_sheet("Best Config")
    sp500_annual = {2020:18.4,2021:28.7,2022:-18.2,2023:26.3,2024:25.0,2025:-2.0}
    rows = [
        ("", "Best Sharpe", "Best Return", "SPY"),
        ("Core N",        best_sharpe["params"]["core_n"],           best_return["params"]["core_n"],           ""),
        ("Momentum",      "189d (9M fixed)",                         "189d (9M fixed)",                         ""),
        ("MA Filter",     best_sharpe["params"]["core_ma"],          best_return["params"]["core_ma"],          ""),
        ("Rebalance",     "Every 42 days",                           "Every 42 days",                           ""),
        ("Sat Threshold", best_sharpe["params"]["sat_score_thresh"], best_return["params"]["sat_score_thresh"], ""),
        ("Sat ATR Stop",  best_sharpe["params"]["sat_stop_atr"],     best_return["params"]["sat_stop_atr"],     ""),
        ("Sat TP",        f"{best_sharpe['params']['sat_take_profit']*100:.0f}%", f"{best_return['params']['sat_take_profit']*100:.0f}%", ""),
        ("Sat Hold",      f"{best_sharpe['params']['sat_hold_days']}d", f"{best_return['params']['sat_hold_days']}d", ""),
        ("Bear Sat",      "Defensive universe", "Defensive universe", ""),
        ("", "", "", ""),
        ("Total Return",  f"{best_sharpe['total_return']}%",  f"{best_return['total_return']}%",  f"{bench_ret:.1f}%"),
        ("CAGR",          f"{best_sharpe['cagr']}%/year",     f"{best_return['cagr']}%/year",     ""),
        ("Win Rate",      f"{best_sharpe['win_rate']}%",      f"{best_return['win_rate']}%",      ""),
        ("Avg Win",       f"{best_sharpe['avg_win']}%",       f"{best_return['avg_win']}%",       ""),
        ("Avg Loss",      f"{best_sharpe['avg_loss']}%",      f"{best_return['avg_loss']}%",      ""),
        ("Profit Factor", best_sharpe["profit_factor"],       best_return["profit_factor"],       ""),
        ("Max Drawdown",  f"{best_sharpe['max_drawdown']}%",  f"{best_return['max_drawdown']}%",  ""),
        ("Sharpe",        best_sharpe["sharpe"],              best_return["sharpe"],              ""),
        ("Total Trades",  best_sharpe["n_trades"],            best_return["n_trades"],            ""),
        ("Core Trades",   best_sharpe["n_core_trades"],       best_return["n_core_trades"],       ""),
        ("Sat Trades",    best_sharpe["n_sat_trades"],        best_return["n_sat_trades"],        ""),
    ]
    for r_idx, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c, value=val)
            if r_idx == 1: cell.fill = hf; cell.font = hft
    for col in "ABCD":
        ws2.column_dimensions[col].width = 25

    ws3 = wb.create_sheet("Best Trades")
    for c, h in enumerate(["Layer","Ticker","Entry","Exit","Entry$","Exit$","P&L%","Reason","Regime","Days"], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hft
    for r_idx, t in enumerate(sorted(best_sharpe["trades"], key=lambda x: x["entry_date"]), 2):
        vals = [t.get("layer",""), t["ticker"], t["entry_date"], t["exit_date"],
                t["entry_price"], t["exit_price"],
                t["pnl_pct"], t["reason"], t.get("regime",""), t["days_held"]]
        fill = gf if t["pnl_pct"] > 0 else rf
        for c, v in enumerate(vals, 1):
            ws3.cell(row=r_idx, column=c, value=v).fill = fill
    for col in "ABCDEFGHIJ":
        ws3.column_dimensions[col].width = 13

    ws4 = wb.create_sheet("Annual Perf")
    for c, h in enumerate(["Year","Model%","SPY%","Trades","WR%"], 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hft
    for r_idx, y in enumerate(best_sharpe.get("annual", []), 2):
        sp = sp500_annual.get(y["year"], "")
        vals = [y["year"], y["return"], sp, y["trades"], y["win_rate"]]
        fill = gf if y["return"] > 0 else rf
        for c, v in enumerate(vals, 1):
            ws4.cell(row=r_idx, column=c, value=v).fill = fill
    for col in "ABCDE":
        ws4.column_dimensions[col].width = 16

    ws5 = wb.create_sheet("Equity Curve")
    for c, h in enumerate(["Date","Model ($)","SPY ($10k)"], 1):
        ws5.cell(row=1, column=c, value=h)
    bench_norm = bench_df["close"]/bench_df["close"].iloc[0]*INITIAL_CASH
    for i, (d, e) in enumerate(zip(all_dates, best_sharpe["equity"]), 2):
        b = bench_norm.get(d)
        ws5.cell(row=i, column=1, value=str(d))
        ws5.cell(row=i, column=2, value=e)
        ws5.cell(row=i, column=3, value=round(float(b),2) if b is not None else None)
    chart = LineChart()
    chart.title = "Combined Model V1 Final vs SPY"
    chart.style = 10
    n_rows = len(best_sharpe["equity"]) + 1
    chart.add_data(Reference(ws5, min_col=2, max_col=3, min_row=1, max_row=n_rows),
                   titles_from_data=True)
    chart.width = 28; chart.height = 16
    ws5.add_chart(chart, "E2")

    wb.save("backtest_results.xlsx")
    logger.info("Saved backtest_results.xlsx")


if __name__ == "__main__":
    run_backtest()